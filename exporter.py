import csv, json
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

class WorkbookExporter:
    """
    This class is responsible for exporting a Workbook instance to various formats.
    """
    def __init__(self, workbook):
        """
        Initializes a new WorkbookExporter instance.
        :param workbook: The Workbook instance to be exported.
        """
        self.workbook = workbook

    # ... other methods ...

    def export_to_excel(self, filename):
        """
        Exports the workbook to an Excel file.
        :param filename: The name of the Excel file to be created.
        """
        for sheet_name, spreadsheet in self.workbook.sheets.items():
            wb = Workbook()
            ws = wb.active
            for i in range(1, spreadsheet.max_row + 1):
                for j in range(1, spreadsheet.max_col_index + 1):
                    cell_value = spreadsheet.get_cell_value(f"{spreadsheet.col_index_to_letter(j)}{i}")
                    ws.cell(row=i, column=j, value=cell_value)
            wb.save(f"{filename}_{sheet_name}.xlsx")


    def export_to_json(self, filename):
        """
        Exports the workbook to a JSON file.
        :param filename: The name of the JSON file to be created.
        """
        for sheet_name, spreadsheet in self.workbook.sheets.items():
            data = {}
            for i in range(1, spreadsheet.max_row + 1):
                for j in range(1, spreadsheet.max_col_index + 1):
                    cell_value = spreadsheet.get_cell_value(f"{spreadsheet.col_index_to_letter(j)}{i}")
                    if i not in data:
                        data[i] = {}
                    data[i][spreadsheet.col_index_to_letter(j)] = cell_value

            with open(f"{filename}_{sheet_name}.json", 'w') as f:
                json.dump(data, f)


    def export_to_pdf(self, filename):
        """
        Exports the workbook to a PDF file.
        :param filename: The name of the PDF file to be created.
        """
        for sheet_name, spreadsheet in self.workbook.sheets.items():
            c = canvas.Canvas(f"{filename}_{sheet_name}.pdf", pagesize=letter)
            width, height = letter
            for i in range(1, spreadsheet.max_row + 1):
                for j in range(1, spreadsheet.max_col_index + 1):
                    cell_value = spreadsheet.get_cell_value(f"{spreadsheet.col_index_to_letter(j)}{i}")
                    c.drawString(10 + j*50, height - i*30, str(cell_value))
            c.save()


    def export_to_csv(self, filename):
        """
        Exports the workbook to a CSV file.
        :param filename: The name of the CSV file to be created.
        """
        for spreadsheet in self.workbook.spreadsheets:
            with open(f"{filename}_{spreadsheet.name}.csv", 'w', newline='') as f:
                writer = csv.writer(f)
                for i in range(1, spreadsheet.max_row + 1):
                    row = [spreadsheet.get_cell_value(f"{spreadsheet.col_index_to_letter(j)}{i}")
                           for j in range(1, spreadsheet.max_col_index + 1)]
                    writer.writerow(row)

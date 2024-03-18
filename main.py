from openpyxl import Workbook, load_workbook
import csv
from fpdf import FPDF
from pathvalidate import validate_filename


class Manager:
    def __init__(self):
        self.filename = None

    def create_workbook(self):
        workbook = Workbook()
        valid_filename = False
        while not valid_filename:
            try:
                self.filename = input("Enter the name of your workbook: ")
                validate_filename(self.filename)
                self.filename = f'{self.filename}.xlsx'
                workbook.active.title = self.filename
                valid_filename = True
            except Exception as e:
                print(e)
        return workbook

    def open_workbook(self):
        filename = input("Enter the filename of the workbook to open: ")
        try:
            workbook = load_workbook(filename)
            print(f"Workbook '{filename}' loaded successfully.")
            return workbook
        except Exception as e:
            print(f"Error: {e}")
            return None

    def get_sheet(self, workbook):
        sheet_name = input("Enter the name of the sheet to open or create: ")
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            workbook.active = workbook[sheet_name]
        return sheet

    def save_workbook(self, workbook):
        workbook.save(f'{workbook.active.title}.xlsx')
        print(f"Workbook saved as {workbook.active.title}.")

    def set_cell_value(self, sheet):
        cell_ref, value = input("Enter cell reference and value (e.g., A1 Hello): ").split(maxsplit=1)
        sheet[cell_ref] = value
        print(f"Cell {cell_ref} set to {value}.")

    def remove_cell_value(self, sheet):
        cell_ref = input("Enter cell reference to remove value (e.g., A1): ")
        if cell_ref in sheet:
            sheet[cell_ref] = None
            print(f"Value removed from cell {cell_ref}.")
        else:
            print(f"Cell {cell_ref} not found.")

    def show_sheet(self, sheet):
        for row in sheet.iter_rows():
            for cell in row:
                print(cell.value, end='\t')
            print()

    def export_csv(self, sheet):
        filename = input("Enter the CSV filename to export: ")
        with open(filename, 'w', newline='') as file:
            writer = csv.writer(file)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
        print(f"Workbook exported as CSV: {filename}.")

    def export_pdf(self, sheet):
        filename = input("Enter the PDF filename to export: ")

        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 12)
                self.cell(0, 10, 'Spreadsheet to PDF', 0, 1, 'C')

            def footer(self):
                self.set_y(-15)
                self.set_font('Arial', 'I', 8)
                self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

        pdf = PDF()
        pdf.add_page()
        pdf.set_font('Arial', '', 12)

        for row in sheet.iter_rows(values_only=True):
            pdf.cell(0, 10, ' '.join(map(str, row)), 0, 1)

        pdf.output(filename)
        print(f"Workbook exported as PDF: {filename}.")


def main():
    manager = Manager()
    choice = input("Do you want to create a new workbook (new) or open an existing one (open)? ").strip().lower()

    if choice == "new":
        workbook = manager.create_workbook()
    elif choice == "open":
        workbook = manager.open_workbook()
        if workbook is None:
            return
    else:
        print("Invalid choice.")
        return

    sheet = manager.get_sheet(workbook)

    while True:
        command = input("Enter a command (set, remove, show, export, save, move, quit): ").strip().lower()

        if command == "set":
            manager.set_cell_value(sheet)

        elif command == "remove":
            manager.remove_cell_value(sheet)

        elif command == "show":
            manager.show_sheet(sheet)

        elif command == "export":
            export_format = input("Enter export format (csv, pdf): ").strip().lower()
            if export_format == "csv":
                manager.export_csv(sheet)
            elif export_format == "pdf":
                manager.export_pdf(sheet)
            else:
                print("Invalid export format.")

        elif command == "save":
            manager.save_workbook(workbook)

        elif command == "move":
            manager.get_sheet(workbook)

        elif command == "quit":
            save_option = input("Do you want to save before quitting? (yes/no): ").strip().lower()
            if save_option == "yes":
                manager.save_workbook(workbook)
            print("Exiting program. Bye!")
            break

        else:
            print("Invalid command. Please try again.")


if __name__ == "__main__":
    main()
